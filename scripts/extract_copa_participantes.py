#!/usr/bin/env python3
"""Extrai participantes da Copa do Brasil (CSV e/ou XLSX) → CSV limpo do Grid.

Fontes (na pasta data/torneios/), mescladas por (ano, nome):
- Copa do Brasil.CSV — dump wiki (vários layouts)
- Copa do Brasil.xlsx — se existir (aba única ano/equipe ou uma aba por ano)
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TORNEIOS = ROOT / "data" / "torneios"
CSV_SRC = TORNEIOS / "Copa do Brasil.CSV"
XLSX_SRC = TORNEIOS / "Copa do Brasil.xlsx"
OUT = TORNEIOS / "participacoes_copa_do_brasil.csv"

EXPECTED: dict[int, int] = {
    **{y: 32 for y in range(1989, 1995)},
    1995: 36,
    1996: 40,
    1997: 44,
    1998: 42,
    1999: 64,
    2000: 69,
    **{y: 64 for y in range(2001, 2013)},
    2013: 87,
    2014: 87,
    2015: 87,
    2016: 86,
    **{y: 91 for y in range(2017, 2021)},
    **{y: 92 for y in range(2021, 2026)},
}

_ESTADOS = (
    r"Acre|Alagoas|Amapá|Amapa|Amazonas|Bahia|Ceará|Ceara|Distrito Federal|"
    r"Espírito Santo|Espirito Santo|Goiás|Goias|Maranhão|Maranhao|"
    r"Mato Grosso do Sul|Mato Grosso|Minas Gerais|Pará|Para|Paraíba|Paraiba|"
    r"Paraná|Parana|Pernambuco|Piauí|Piaui|Rio de Janeiro|Rio Grande do Norte|"
    r"Rio Grande do Sul|Rondônia|Rondonia|Roraima|Santa Catarina|"
    r"São Paulo|Sao Paulo|Sergipe|Tocantins"
)
_ESTADO_RE = re.compile(rf"^(?:{_ESTADOS})\b", re.I)
_REGIAO_RE = re.compile(
    r"^Região\s+(Norte|Nordeste|Centro-Oeste|Sudeste|Sul)\b", re.I
)

_SKIP_EXACT = {
    "estatísticas",
    "estatisticas",
    "colocações finais",
    "colocacoes finais",
    "melhor marcador",
    "maior goleada",
    "maiores goleadas",
    "(diferença)",
    "(diferenca)",
    "campeão",
    "campeao",
    "vice-campeão",
    "vice-campeao",
    "posição",
    "posicao",
    "pos.",
    "pontos",
    "clube",
    "equipe",
    "estado",
    "uf",
    "competição",
    "competicao",
    "forma de classificação",
    "forma de classificacao",
    "forma de entrada",
    "como se classificou",
    "classificados por outras competições",
    "classificados por outras competicoes",
}
_SKIP_SUB = (
    "como se class",
    "forma de class",
    "forma de entrada",
    "regi",
    "ranking",
    "federação",
    "federacao",
    "estadual",
    "metropolitano",
    "colocações",
    "colocacoes",
    "classificados por",
    "colocado",
    "primeira fase",
    "segunda fase",
    "2ª fase",
    "por partida",
    "estádio",
    "estadio",
    "fevereiro",
    "março",
    "marco",
    "abril",
    "maio",
    "junho",
    "julho",
    "agosto",
    "setembro",
    "outubro",
    "novembro",
    "dezembro",
)

_FORM_LABELS = {
    "como se classificou",
    "como se classificou",
    "forma de classificação",
    "forma de classificacao",
    "forma de entrada",
}
_RANK_LABELS = {"posição", "posicao", "pos.", "pos", "pontos"}


def clean_name(s: str) -> str:
    s = (s or "").replace("\xa0", " ").strip()
    s = re.sub(r"\[[^\]]*\]", "", s)
    s = re.sub(r"\s*\([^)]*t[ií]tulo[^)]*\)\s*", " ", s, flags=re.I)
    s = re.sub(r"\s*\(\d+\.?\º?\s*t[ií]tulo\)\s*", " ", s, flags=re.I)
    s = re.sub(r"\s*\([^)]*campe[aã]o[^)]*\)\s*", " ", s, flags=re.I)
    s = re.sub(r"\s*\([^)]*vice[^)]*\)\s*", " ", s, flags=re.I)
    s = re.sub(r"\s*\([^)]*apenas[^)]*\)\s*", " ", s, flags=re.I)
    s = re.sub(r"\s*\(nota:[^)]*\)\s*", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip(" .;")


def standalone_year(cell: object) -> int | None:
    if isinstance(cell, int) and 1980 <= cell <= 2035:
        return cell
    if isinstance(cell, float) and cell == int(cell) and 1980 <= int(cell) <= 2035:
        return int(cell)
    c = str(cell or "").replace("\xa0", " ").strip()
    return int(c) if re.fullmatch(r"(19|20)\d{2}", c) else None


def is_estado_header(nome: str) -> bool:
    n = clean_name(nome)
    if not n:
        return False
    if _REGIAO_RE.match(n):
        return True
    # "Acre (AC)", "Bahia (BA)", "São Paulo (7): ..."
    if re.fullmatch(rf"(?:{_ESTADOS})\s*(\([A-Z]{{2}}\))?", n, flags=re.I):
        return True
    if re.match(rf"^(?:{_ESTADOS})\s*\(\d+\)\s*:", n, flags=re.I):
        return False  # region list with count — not a bare header
    if re.fullmatch(rf"(?:{_ESTADOS})\s*\([^)]*\)", n, flags=re.I):
        return True
    return False


def is_club_name(nome: str) -> bool:
    if not nome or len(nome) < 2:
        return False
    low = nome.casefold()
    if low in _SKIP_EXACT:
        return False
    if any(x in low for x in _SKIP_SUB):
        return False
    if re.fullmatch(r"[A-Z]{2}", nome):
        return False
    if re.fullmatch(r"[\d\s.º°oª]+", nome):
        return False
    if low.startswith("campe") or low.startswith("vice"):
        return False
    if is_estado_header(nome):
        return False
    if _REGIAO_RE.match(nome):
        return False
    # Placar / artilheiro / datas misturados no dump
    if re.search(r"\d\s*[–\-xX×]\s*\d", nome):
        return False
    if "gol" in low:
        return False
    if re.match(r"^\d", nome):
        return False
    # "André Lima (Botafogo)" — jogador; permite só sufixo de UF "(SP)"
    m_par = re.search(r"\(([^)]+)\)\s*$", nome)
    if m_par and not re.fullmatch(r"[A-Z]{2}", m_par.group(1).strip()):
        return False
    return True


def split_region_clubs(blob: str) -> list[str]:
    blob = clean_name(blob)
    blob = re.sub(rf"^(?:{_ESTADOS})\s*(\([^)]*\))?\s*:\s*", "", blob, flags=re.I)
    parts = re.split(r",\s*|\s+e\s+", blob)
    out: list[str] = []
    for p in parts:
        p = clean_name(p)
        if is_club_name(p) and not re.fullmatch(_ESTADOS, p, flags=re.I):
            out.append(p)
    return out


def _add(bucket: dict[int, set[str]], ano: int, nome: str) -> None:
    nome = clean_name(nome)
    if 1989 <= ano <= 2030 and is_club_name(nome):
        bucket[ano].add(nome)


def _read_csv_rows(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="latin-1")
    return [list(csv.reader([ln], delimiter=";"))[0] for ln in text.splitlines()]


def _lab(cell: str) -> str:
    return (cell or "").replace("\xa0", " ").strip().casefold()


def _is_club_header(lab: str) -> bool:
    return lab in ("equipe", "clube", "time", "times")


def _col_kind(hdr: list[str], i: int) -> str:
    """'part' (participantes), 'rank' (ranking CBF) ou 'unk'."""
    neighbors = []
    for j in (i - 1, i + 1, i + 2):
        if 0 <= j < len(hdr):
            neighbors.append(_lab(hdr[j]))
    if any(n in _FORM_LABELS for n in neighbors):
        return "part"
    if any(n in _RANK_LABELS for n in neighbors):
        return "rank"
    # UF/Estado imediatamente à esquerda costuma ser tabela de participantes
    if i > 0 and _lab(hdr[i - 1]) in ("estado", "uf"):
        return "part"
    return "unk"


def extract_table_columns(rows: list[list[str]]) -> dict[int, set[str]]:
    """Extrai clubes de colunas Equipe/Clube com preferência por participantes."""
    out: dict[int, set[str]] = defaultdict(set)
    header_idxs = [
        i
        for i, r in enumerate(rows)
        if any(_is_club_header(_lab(c)) for c in r)
    ]

    # ano -> lista de (score, n, col_idx, set)
    candidatos: dict[int, list[tuple[float, int, int, set[str]]]] = defaultdict(list)

    for hi in header_idxs:
        hdr = rows[hi]
        year_map: dict[int, int] = {}
        for back in range(1, 8):
            if hi - back < 0:
                break
            yr = rows[hi - back]
            hits = [(j, standalone_year(c)) for j, c in enumerate(yr) if standalone_year(c)]
            if not hits:
                if year_map:
                    break
                continue
            for j, ano in hits:
                assert ano is not None
                year_map.setdefault(j, ano)

        if not year_map:
            continue

        eq_cols: list[tuple[int, int, str]] = []
        for i, c in enumerate(hdr):
            if not _is_club_header(_lab(c)):
                continue
            ano = None
            for j in range(i, -1, -1):
                if j in year_map:
                    ano = year_map[j]
                    break
            if ano:
                eq_cols.append((i, ano, _col_kind(hdr, i)))

        if not eq_cols:
            continue

        end = len(rows)
        for nxt in header_idxs:
            if nxt > hi:
                end = nxt
                break
        for r_i in range(hi + 1, end):
            if sum(1 for c in rows[r_i] if standalone_year(c)) >= 2:
                end = r_i
                break

        col_sets: dict[int, set[str]] = defaultdict(set)
        for r in rows[hi + 1 : end]:
            for i, ano, _kind in eq_cols:
                if i < len(r):
                    nome = clean_name(r[i] or "")
                    if is_club_name(nome):
                        col_sets[i].add(nome)

        for i, ano, kind in eq_cols:
            clubs = col_sets.get(i) or set()
            if not clubs:
                continue
            n = len(clubs)
            exp = EXPECTED.get(ano)
            dist = abs(n - exp) if exp else 0
            # Ranking costuma ter ~10–20; penaliza forte
            penalty = 0.0
            if kind == "rank":
                penalty += 40
            elif kind == "part":
                penalty -= 5
            if exp and n < exp * 0.45:
                penalty += 25
            if exp and n > exp * 1.35:
                penalty += 30
            score = dist + penalty
            candidatos[ano].append((score, n, i, clubs))

    for ano, opts in candidatos.items():
        opts.sort(key=lambda t: (t[0], -t[1]))
        best = opts[0][3]
        out[ano] |= best
    return out


def extract_block_lists(rows: list[list[str]]) -> dict[int, set[str]]:
    """Listas por região / UF e colunas soltas de clubes sob banners de ano."""
    out: dict[int, set[str]] = defaultdict(set)

    banner_idxs = [
        i
        for i, r in enumerate(rows)
        if sum(1 for c in r if standalone_year(c)) >= 2
    ]

    for bi, i in enumerate(banner_idxs):
        r = rows[i]
        years_here = [(j, standalone_year(c)) for j, c in enumerate(r) if standalone_year(c)]
        bounds: list[tuple[int, int, int]] = []
        for k, (j, ano) in enumerate(years_here):
            assert ano is not None
            j2 = years_here[k + 1][0] if k + 1 < len(years_here) else max(len(r), 40)
            bounds.append((ano, j, j2))

        end = banner_idxs[bi + 1] if bi + 1 < len(banner_idxs) else len(rows)
        # Para no próximo banner; também corta se aparecer outro bloco multi-ano

        for r2 in rows[i + 1 : end]:
            if sum(1 for c in r2 if standalone_year(c)) >= 2:
                break
            for ano, c0, c1 in bounds:
                for j in range(c0, min(c1, len(r2))):
                    cell = (r2[j] or "").replace("\xa0", " ").strip()
                    if not cell:
                        continue
                    # "São Paulo: Corinthians, …" / "Bahia (2): Bahia e Vitória."
                    if ":" in cell and (_ESTADO_RE.match(cell) or _REGIAO_RE.match(cell)):
                        for nome in split_region_clubs(cell):
                            _add(out, ano, nome)
                        continue
                    if is_estado_header(cell):
                        continue
                    _add(out, ano, cell)
    return out


def merge_prefer_tables(
    tables: dict[int, set[str]],
    blocks: dict[int, set[str]],
) -> dict[int, set[str]]:
    """Usa tabelas Equipe/Clube; listas regionais completam anos ainda abaixo do esperado."""
    out: dict[int, set[str]] = defaultdict(set)
    for ano, clubs in tables.items():
        out[ano] |= clubs
    for ano, clubs in blocks.items():
        exp = EXPECTED.get(ano)
        atual = len(out.get(ano, ()))
        if exp and atual >= exp - 4:
            continue
        for nome in clubs:
            if not is_club_name(nome):
                continue
            out[ano].add(nome)
            if exp and len(out[ano]) >= exp + 8:
                break
    return out


def extract_from_csv(path: Path) -> dict[int, set[str]]:
    rows = _read_csv_rows(path)
    return merge_prefer_tables(extract_table_columns(rows), extract_block_lists(rows))


def extract_from_xlsx(path: Path) -> dict[int, set[str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl necessário para ler o XLSX") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[int, set[str]] = defaultdict(set)
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheet_year = standalone_year(sheet)
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = None
            header_i = 0
            for i, row in enumerate(rows[:12]):
                cells = [
                    str(c).casefold().strip() if c is not None else "" for c in row
                ]
                if any(c in ("equipe", "clube", "time", "times") for c in cells) and (
                    any(c in ("ano", "edição", "edicao", "season") for c in cells)
                    or sheet_year
                ):
                    header = cells
                    header_i = i
                    break
            if header:
                idx_ano = next(
                    (
                        i
                        for i, c in enumerate(header)
                        if c in ("ano", "edição", "edicao", "season")
                    ),
                    None,
                )
                idx_eq = next(
                    (
                        i
                        for i, c in enumerate(header)
                        if c in ("equipe", "clube", "time", "times", "nome")
                    ),
                    None,
                )
                if idx_eq is None:
                    continue
                for row in rows[header_i + 1 :]:
                    if not row:
                        continue
                    ano = sheet_year
                    if idx_ano is not None and idx_ano < len(row):
                        ano = standalone_year(row[idx_ano]) or ano
                    if ano is None:
                        continue
                    if idx_eq < len(row):
                        _add(out, ano, str(row[idx_eq] or ""))
                continue
            if sheet_year:
                for row in rows:
                    if not row:
                        continue
                    for cell in row[:3]:
                        nome = clean_name(str(cell or ""))
                        if is_club_name(nome) and not standalone_year(cell):
                            _add(out, sheet_year, nome)
                            break
    finally:
        wb.close()
    return out


def merge_buckets(*buckets: dict[int, set[str]]) -> dict[int, set[str]]:
    out: dict[int, set[str]] = defaultdict(set)
    for b in buckets:
        for ano, nomes in b.items():
            out[ano] |= nomes
    return out


def write_csv(bucket: dict[int, set[str]], path: Path) -> int:
    rows: list[dict[str, str]] = []
    for ano in sorted(bucket):
        for nome in sorted(bucket[ano], key=lambda s: s.casefold()):
            rows.append(
                {
                    "competicao": "copa_do_brasil",
                    "ano": str(ano),
                    "nome": nome,
                }
            )
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["competicao", "ano", "nome"], delimiter=";")
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def main() -> None:
    buckets: list[dict[int, set[str]]] = []
    if CSV_SRC.is_file():
        b = extract_from_csv(CSV_SRC)
        print(f"CSV: {sum(len(v) for v in b.values())} entradas em {len(b)} anos")
        buckets.append(b)
    else:
        print(f"CSV ausente: {CSV_SRC.name}")

    if XLSX_SRC.is_file():
        b = extract_from_xlsx(XLSX_SRC)
        print(f"XLSX: {sum(len(v) for v in b.values())} entradas em {len(b)} anos")
        buckets.append(b)
    else:
        print(f"XLSX ausente (ok se só houver CSV): {XLSX_SRC.name}")

    if not buckets:
        raise SystemExit("Nenhuma fonte de participantes da Copa encontrada")

    merged = merge_buckets(*buckets)
    n = write_csv(merged, OUT)
    print(f"→ {OUT.name}: {n} linhas")
    ok = baixo = 0
    for ano in sorted(EXPECTED):
        got = len(merged.get(ano, ()))
        exp = EXPECTED[ano]
        # Margem: ≥ ~70% do esperado e ≥ 20 clubes (ou ≥ esperado-12)
        bom = got >= max(20, int(exp * 0.70)) or got >= exp - 12
        flag = "ok" if bom else "BAIXO"
        if bom:
            ok += 1
        else:
            baixo += 1
        print(f"  {ano}: {got:3d}/{exp} {flag}")
    print(f"Resumo: {ok} ok, {baixo} BAIXO")


if __name__ == "__main__":
    main()
