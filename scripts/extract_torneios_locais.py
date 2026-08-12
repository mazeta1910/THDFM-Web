#!/usr/bin/env python3
"""Extrai dados do Grid a partir dos arquivos locais em data/torneios/ (sem scrape).

Fontes:
- Brasileirao Serie B.CSV → campeão/vice/3º/4º (linhas Detalhes)
- Série C.xlsx / Serie C.CSV → classificações finais (via extract_serie_c_classif)
- Copa do Brasil.CSV → campeões e vices
- Goleadas.xlsx → goleadas Série A/B e Copa do Brasil
"""

from __future__ import annotations

import csv
import re
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TORNEIOS = ROOT / "data" / "torneios"

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


def _clean(s: str) -> str:
    s = (s or "").replace("\xa0", " ").strip()
    s = re.sub(r"\[[^\]]*\]", "", s)
    s = re.sub(r"\s*\([^)]*t[ií]tulo[^)]*\)\s*", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


def extract_serie_b_top4() -> Path:
    """Linhas Detalhes do dump wiki local → posições 1–4 por ano."""
    src = TORNEIOS / "Brasileirao Serie B.CSV"
    lines = src.read_text(encoding="latin-1").splitlines()
    rows: list[dict] = []
    for i, line in enumerate(lines):
        if not re.match(r"^(19|20)\d{2};", line):
            continue
        if "Não disputado" in line:
            continue
        ano = int(line.split(";", 1)[0][:4])
        detalhes = None
        for j in range(i + 1, min(i + 4, len(lines))):
            cells = lines[j].split(";")
            if cells and cells[0].strip() == "Detalhes":
                detalhes = cells
                break
        if not detalhes or len(detalhes) < 6:
            continue
        nomes = [_clean(detalhes[k]) for k in (1, 3, 4, 5)]
        if not nomes[0]:
            continue
        for pos, nome in enumerate(nomes, start=1):
            if not nome:
                continue
            rows.append(
                {
                    "competicao": "serie_b",
                    "ano": ano,
                    "posicao": pos,
                    "n_clubes": 4,
                    "nome": nome,
                    "pts": "",
                    "j": "",
                    "v": "",
                    "e": "",
                    "d": "",
                    "gp": "",
                    "gc": "",
                    "sg": "",
                    "fonte_url": "",
                }
            )
    out = TORNEIOS / "classificacoes_serie_b.csv"
    _write(out, rows)
    print(f"Série B top4: {len(rows)} linhas, {len({r['ano'] for r in rows})} anos → {out.name}")
    return out


def extract_copa_campeoes() -> Path:
    src = TORNEIOS / "Copa do Brasil.CSV"
    text = src.read_text(encoding="latin-1")
    camps: list[str] = []
    vices: list[str] = []
    for m in re.finditer(r"(?:^|;)Campeão;([^;]+)", text):
        nome = _clean(m.group(1))
        if not nome or "estadual" in nome.casefold() or "coloc" in nome.casefold():
            continue
        camps.append(nome)
    for m in re.finditer(r"(?:^|;)Vice-campeão;([^;]+)", text):
        nome = _clean(m.group(1))
        if not nome or "estadual" in nome.casefold():
            continue
        vices.append(nome)
    # Anos 1989… na ordem dos blocos
    n = min(len(camps), len(vices))
    rows = [
        {
            "competicao": "copa_do_brasil",
            "ano": 1989 + i,
            "campeao": camps[i],
            "vice": vices[i],
        }
        for i in range(n)
    ]
    out = TORNEIOS / "campeoes_copa_do_brasil.csv"
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f, fieldnames=["competicao", "ano", "campeao", "vice"], delimiter=";"
        )
        w.writeheader()
        w.writerows(rows)
    print(f"Copa campeões: {len(rows)} edições → {out.name}")
    print(f"  {rows[0]} … {rows[-1]}")
    return out


def extract_goleadas() -> None:
    if openpyxl is None:
        raise SystemExit("openpyxl necessário só para exportar Goleadas.xlsx")
    wb = openpyxl.load_workbook(TORNEIOS / "Goleadas.xlsx", read_only=True, data_only=True)

    liga_rows: list[dict] = []
    for i, r in enumerate(wb["Ligas"].iter_rows(values_only=True)):
        if i == 0 or not r or r[0] not in ("serie_a", "serie_b", "serie_c"):
            continue
        comp, ano, rank, m, v, gm, gv = (list(r) + [None] * 7)[:7]
        try:
            gm_i, gv_i, ano_i = int(gm), int(gv), int(ano)
        except Exception:
            continue
        if gm_i > gv_i:
            win, lose, diff = m, v, gm_i - gv_i
        elif gv_i > gm_i:
            win, lose, diff = v, m, gv_i - gm_i
        else:
            continue
        liga_rows.append(
            {
                "competicao": comp,
                "ano": ano_i,
                "rank": int(rank or 1),
                "mandante": _clean(str(m or "")),
                "visitante": _clean(str(v or "")),
                "gols_mandante": gm_i,
                "gols_visitante": gv_i,
                "vencedor": _clean(str(win or "")),
                "perdedor": _clean(str(lose or "")),
                "diff": diff,
                "obs": _clean(str(r[10] or "")) if len(r) > 10 else "",
            }
        )
    liga_rows.sort(key=lambda x: (x["competicao"], x["ano"], -x["diff"]))
    _write_dict(TORNEIOS / "goleadas_ligas.csv", liga_rows)
    a = [r for r in liga_rows if r["competicao"] == "serie_a"]
    _write_dict(TORNEIOS / "goleadas_serie_a.csv", a)
    print(f"Goleadas ligas: {len(liga_rows)} (A={len(a)})")

    copa: list[dict] = []
    for i, r in enumerate(wb["Copa do Brasil"].iter_rows(values_only=True)):
        if i == 0 or not r or not r[1]:
            continue
        mand = _clean(str(r[1] or ""))
        placar = str(r[2] or "").replace("\xa0", " ").strip()
        visit = _clean(str(r[3] or ""))
        m = re.match(r"(\d+)\s*[–\-xX]\s*(\d+)", placar)
        if not m:
            continue
        gm, gv = int(m.group(1)), int(m.group(2))
        if gm > gv:
            win, lose, diff = mand, visit, gm - gv
        elif gv > gm:
            win, lose, diff = visit, mand, gv - gm
        else:
            continue
        try:
            ano_i = int(r[6])
        except Exception:
            ano_i = ""
        copa.append(
            {
                "competicao": "copa_do_brasil",
                "ano": ano_i,
                "mandante": mand,
                "visitante": visit,
                "gols_mandante": gm,
                "gols_visitante": gv,
                "vencedor": win,
                "perdedor": lose,
                "diff": diff,
                "placar": placar,
                "estadio": _clean(str(r[4] or "")),
                "data": _clean(str(r[5] or "")),
            }
        )
    wb.close()
    _write_dict(TORNEIOS / "goleadas_copa_do_brasil.csv", copa)
    print(f"Goleadas Copa: {len(copa)}")


def _write(path: Path, rows: list[dict]) -> None:
    fields = [
        "competicao",
        "ano",
        "posicao",
        "n_clubes",
        "nome",
        "pts",
        "j",
        "v",
        "e",
        "d",
        "gp",
        "gc",
        "sg",
        "fonte_url",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        w.writerows(rows)


def _write_dict(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=";")
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    extract_serie_b_top4()
    # Série C: reusa extrator do xlsx local
    c_script = ROOT / "scripts" / "extract_serie_c_classif.py"
    if c_script.is_file():
        subprocess.check_call([sys.executable, str(c_script)])
    extract_copa_campeoes()
    extract_goleadas()
    print("OK — tudo a partir de data/torneios/ local (sem Wikipedia).")


if __name__ == "__main__":
    main()
