#!/usr/bin/env python3
"""Extrai classificacao completa da Serie B a partir de Serie B.xlsx.

Aba unica 'Brasileirao Serie B' com blocos por ano (1971+) lado a lado.
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "torneios" / "Serie B.xlsx"
SHEET = "Brasileirao Serie B"
OUT = ROOT / "data" / "torneios" / "classificacoes_serie_b.csv"
FIELDNAMES = [
    "competicao",
    "ano",
    "posicao",
    "n_clubes",
    "nome",
    "pts",
    "j",
    "v",
    "e",
    "d",
    "gp",
    "gc",
    "sg",
    "fonte_url",
]

HEADER_CLUB = {"time", "equipe", "equipes", "clube", "team", "club"}
HEADER_PTS = {"pts", "pontos", "pg", "p", "points"}

STOP_CLUB = {
    "clube",
    "time",
    "equipe",
    "equipes",
    "team",
    "club",
    "pos",
    "campeão",
    "campeao",
    "champions",
    "champion",
    "vice-campeão",
    "vice-campeao",
    "promovido(s)",
    "promoted",
    "rebaixado(s)",
    "relegated",
    "estatísticas",
    "estatisticas",
    "colocações finais",
    "colocacoes finais",
    "melhor marcador",
    "melhor ataque",
    "melhor defesa",
    "maior goleada",
    "top goalscorer",
    "source",
    "notes",
    "(diferença)",
    "(diferenca)",
    "tabela de classificação",
    "tabela de classificacao",
}


def _norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int) and not isinstance(v, bool):
        return str(v)
    s = str(v).replace("\xa0", " ").strip()
    return s.replace("–", "-").replace("—", "-").replace("−", "-")


def _as_int(s: str) -> int | None:
    s = _norm(s)
    m = re.match(r"^(-?\d+)", s)
    return int(m.group(1)) if m else None


def _is_year(s: str) -> bool:
    return bool(re.fullmatch(r"(19|20)\d{2}", s)) and 1971 <= int(s) <= 2030


def _looks_club(name: str) -> bool:
    if not name or len(name) < 2:
        return False
    low = name.lower().strip()
    if low in STOP_CLUB:
        return False
    if low.startswith("source") or low.startswith("group ") or low.startswith("notes"):
        return False
    if re.fullmatch(r"\d+", name):
        return False
    return bool(re.search(r"[A-Za-zÀ-ÿ]", name))


def _clean_club(name: str) -> str:
    name = re.sub(r"\[[^\]]*\]", "", name).strip()
    name = re.sub(r"\s*\(C\)\s*$", "", name, flags=re.I).strip()
    name = re.sub(r"\s*\(.*?\)\s*$", "", name).strip()
    name = re.sub(r"\d+$", "", name).strip()
    return name


def _fmt_sg(sg: int | None) -> str:
    if sg is None:
        return ""
    return f"+{sg}" if sg > 0 else str(sg)


def _load_grid(path: Path) -> list[list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if SHEET not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"Aba '{SHEET}' nao encontrada em {path.name}")
    ws = wb[SHEET]
    grid = [
        [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]
    wb.close()
    return grid


def _col_end(grid: list[list[str]], yr: int, yc: int, year_markers: list[tuple[int, int, int]]) -> int:
    """Limite direito do bloco = proximo ano na mesma linha, senao yc+14."""
    same_row = sorted(c for r, c, _a in year_markers if r == yr and c > yc)
    if same_row:
        return same_row[0]
    # Anos em linhas proximas (±2) a direita
    near = sorted(
        c
        for r, c, _a in year_markers
        if abs(r - yr) <= 2 and c > yc
    )
    if near:
        return min(near[0], yc + 16)
    return min(len(grid[yr]), yc + 14)


def _map_header_row(
    row: list[str], c0: int, c1: int
) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    texts = [t.lower() for t in row]
    english = any(
        texts[j] in {"team", "pld", "pts", "gf", "ga", "gd", "points"}
        for j in range(c0, min(c1, len(texts)))
        if texts[j]
    )
    for j in range(c0, min(c1, len(row))):
        t = texts[j]
        if not t:
            continue
        if t in {"#", "pos", "pos.", "posição", "posicao", "col.", "col", "class."} or (
            t.startswith("pos") and t != "possession"
        ):
            mapping.setdefault("posicao", j)
        elif t in HEADER_CLUB:
            mapping.setdefault("clube", j)
        elif t in HEADER_PTS or t.startswith("pt"):
            mapping.setdefault("pts", j)
        elif t in {"j", "jogos", "pj", "pld", "apps"}:
            mapping.setdefault("j", j)
        elif t in {"v", "vit", "vitórias", "vitorias", "w", "win", "wins"}:
            mapping.setdefault("v", j)
        elif english and t in {"d", "draw", "draws"}:
            mapping.setdefault("e", j)
        elif (not english) and t in {"e", "emp", "empates"}:
            mapping.setdefault("e", j)
        elif english and t in {"l", "loss", "losses"}:
            mapping.setdefault("d", j)
        elif (not english) and t in {"d", "der", "derrotas"}:
            mapping.setdefault("d", j)
        elif t in {"gp", "gf", "gols pró", "gols pro"}:
            mapping.setdefault("gp", j)
        elif t in {"gc", "ga", "gols contra"}:
            mapping.setdefault("gc", j)
        elif t in {"sg", "saldo", "gd"}:
            mapping.setdefault("sg", j)
    if "pts" in mapping and "clube" in mapping:
        return mapping
    return None


def _parse_rows_after(
    grid: list[list[str]], hr: int, mapping: dict[str, int], c0: int, c1: int
) -> list[dict]:
    out: list[dict] = []
    for r in range(hr + 1, min(hr + 55, len(grid))):
        row = grid[r]
        # Novo ano nesta faixa? encerra
        for c in range(c0, min(c1, len(row))):
            if _is_year(row[c]) and r > hr:
                return out
        ci = mapping["clube"]
        clube = row[ci] if ci < len(row) else ""
        if not _looks_club(clube):
            alt = row[ci + 1] if ci + 1 < len(row) and ci + 1 < c1 else ""
            if _as_int(clube) is not None and _looks_club(alt):
                clube = alt
            else:
                if out:
                    break
                continue
        pts = _as_int(row[mapping["pts"]]) if mapping["pts"] < len(row) else None
        if pts is None:
            if out:
                break
            continue

        def grab(key: str) -> int | None:
            j = mapping.get(key)
            if j is None or j >= len(row) or j >= c1:
                return None
            return _as_int(row[j])

        pos = grab("posicao")
        if pos is None and ci > c0:
            pos = _as_int(row[ci - 1])
        if pos is None:
            pos = len(out) + 1

        jg, v, e, d = grab("j"), grab("v"), grab("e"), grab("d")
        gp, gc, sg = grab("gp"), grab("gc"), grab("sg")
        if sg is None and gp is not None and gc is not None:
            sg = gp - gc
        if jg is None and None not in (v, e, d):
            jg = (v or 0) + (e or 0) + (d or 0)

        out.append(
            {
                "posicao": pos,
                "nome": _clean_club(clube),
                "pts": pts,
                "j": jg,
                "v": v,
                "e": e,
                "d": d,
                "gp": gp,
                "gc": gc,
                "sg": sg,
            }
        )
        if len(out) >= 48:
            break
    return out


def _find_campeao(
    grid: list[list[str]], yr: int, c0: int, c1: int, row_limit: int
) -> str | None:
    """Campeão só no entorno desta temporada (nao vazar para bloco abaixo)."""
    for r in range(yr, min(row_limit, len(grid))):
        if r > yr:
            for c in range(c0, min(c1, len(grid[r]))):
                if _is_year(grid[r][c]):
                    return None
        for c in range(c0, min(c1, len(grid[r]))):
            cell = grid[r][c].lower()
            if cell not in {"campeão", "campeao", "champions", "champion"}:
                continue
            for cc in (c + 1, c + 2):
                if cc < c1 and cc < len(grid[r]) and _looks_club(grid[r][cc]):
                    return _clean_club(grid[r][cc])
            if r + 1 < len(grid):
                if c < len(grid[r + 1]) and c < c1 and _looks_club(grid[r + 1][c]):
                    return _clean_club(grid[r + 1][c])
                if (
                    c + 1 < len(grid[r + 1])
                    and c + 1 < c1
                    and _looks_club(grid[r + 1][c + 1])
                ):
                    return _clean_club(grid[r + 1][c + 1])
    return None


def _apply_campeao(rows: list[dict], campeao: str | None) -> list[dict]:
    if not rows or not campeao:
        return rows
    key = campeao.casefold()
    by = {r["nome"].casefold(): r for r in rows}
    if key not in by:
        rows = [
            {
                "posicao": 1,
                "nome": campeao,
                "pts": None,
                "j": None,
                "v": None,
                "e": None,
                "d": None,
                "gp": None,
                "gc": None,
                "sg": None,
            }
        ] + rows
    elif rows[0]["nome"].casefold() != key:
        champ = by[key]
        others = [r for r in rows if r["nome"].casefold() != key]
        rows = [champ] + others
    for i, r in enumerate(rows, start=1):
        r["posicao"] = i
    return rows


def _dedupe_keep_order(rows: list[dict]) -> list[dict]:
    seen: set[str] = set()
    out: list[dict] = []
    for r in rows:
        k = r["nome"].casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(r)
    for i, r in enumerate(out, start=1):
        r["posicao"] = i
    return out


def extract_rows(path: Path = XLSX) -> list[dict]:
    grid = _load_grid(path)
    year_markers: list[tuple[int, int, int]] = []
    for r, row in enumerate(grid):
        for c, val in enumerate(row):
            if _is_year(val):
                year_markers.append((r, c, int(val)))

    by_ano: dict[int, list[dict]] = {}
    for yr, yc, ano in year_markers:
        if ano > 2025:
            continue
        c0 = max(0, yc - 1)
        c1 = _col_end(grid, yr, yc, year_markers)

        # Primeiro cabecalho apos o ano = tabela desta temporada.
        # Cabecalhos seguintes na mesma faixa costumam ser de outros anos
        # empilhados mais abaixo — so aceitamos extras se forem grupos EN
        # logo abaixo (bloco ainda curto).
        first_hdr: tuple[int, dict[str, int]] | None = None
        extra_hdrs: list[tuple[int, dict[str, int]]] = []
        for r in range(yr, min(yr + 55, len(grid))):
            if r > yr and yc < len(grid[r]) and _is_year(grid[r][yc]):
                break
            # Ano vizinho nesta faixa de colunas
            if r > yr + 1:
                hit_other = False
                for c in range(c0, min(c1, len(grid[r]))):
                    if c != yc and _is_year(grid[r][c]):
                        hit_other = True
                        break
                if hit_other and first_hdr is not None:
                    break
            mapping = _map_header_row(grid[r], c0, c1)
            if not mapping:
                continue
            if first_hdr is None:
                first_hdr = (r, mapping)
            else:
                # Grupos adicionais so a ate ~25 linhas do primeiro cabecalho
                if r - first_hdr[0] <= 30:
                    extra_hdrs.append((r, mapping))
                else:
                    break

        if first_hdr is None:
            continue

        main = _parse_rows_after(grid, first_hdr[0], first_hdr[1], c0, c1)
        if not main:
            continue

        merged = list(main)
        seen = {r["nome"].casefold() for r in merged}
        # So mescla grupos extras se a tabela principal parece fase de grupo (< 12).
        if len(main) < 12:
            for hr, mp in extra_hdrs:
                for row in _parse_rows_after(grid, hr, mp, c0, c1):
                    k = row["nome"].casefold()
                    if k not in seen:
                        seen.add(k)
                        nr = dict(row)
                        nr["posicao"] = len(merged) + 1
                        merged.append(nr)

        merged = _dedupe_keep_order(merged)
        # Janela do campeao: do ano ate pouco depois do fim da tabela
        row_limit = first_hdr[0] + len(main) + 20
        campeao = _find_campeao(grid, yr, c0, c1, row_limit)
        merged = _apply_campeao(merged, campeao)
        if len(merged) < 4:
            continue

        prev = by_ano.get(ano)
        if prev is None or len(merged) > len(prev):
            by_ano[ano] = merged

    rows: list[dict] = []
    for ano in sorted(by_ano):
        block = by_ano[ano]
        n = len(block)
        for r in block:
            sg = r["sg"]
            rows.append(
                {
                    "competicao": "serie_b",
                    "ano": ano,
                    "posicao": r["posicao"],
                    "n_clubes": n,
                    "nome": r["nome"],
                    "pts": r["pts"] if r["pts"] is not None else "",
                    "j": r["j"] if r["j"] is not None else "",
                    "v": r["v"] if r["v"] is not None else "",
                    "e": r["e"] if r["e"] is not None else "",
                    "d": r["d"] if r["d"] is not None else "",
                    "gp": r["gp"] if r["gp"] is not None else "",
                    "gc": r["gc"] if r["gc"] is not None else "",
                    "sg": _fmt_sg(sg) if isinstance(sg, int) else "",
                    "fonte_url": "",
                }
            )
    return rows


def main() -> int:
    if not XLSX.is_file():
        print(f"Arquivo nao encontrado: {XLSX}", file=sys.stderr)
        return 1
    rows = extract_rows()
    if not rows:
        print("Nenhuma tabela extraida do xlsx.", file=sys.stderr)
        return 1
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES, delimiter=";")
        w.writeheader()
        w.writerows(rows)
    anos = sorted({int(r["ano"]) for r in rows})
    print(
        f"OK {OUT.relative_to(ROOT)}: {len(rows)} linhas, "
        f"anos {anos[0]}-{anos[-1]} ({len(anos)} temporadas)"
    )
    # Anos sem disputa da Série B (não é lacuna do xlsx).
    sem_disputa = {1973, 1974, 1975, 1976, 1977, 1978, 1979, 1993, 2000}
    faltando = [
        a for a in range(1971, 2026) if a not in set(anos) and a not in sem_disputa
    ]
    if faltando:
        print(f"  lacunas inesperadas no xlsx: {faltando}")
    else:
        print(f"  cobertura completa (exceto anos sem disputa: {sorted(sem_disputa)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
